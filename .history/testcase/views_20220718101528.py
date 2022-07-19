from django.contrib import messages
from django.http import HttpResponseRedirect,HttpResponse,JsonResponse
from django.shortcuts import render,redirect
from django.urls import reverse
from django.utils.translation import gettext_lazy as _
import pandas as pd
from django.forms import inlineformset_factory
from testcase.forms import (
    SearchCaseFormTest,
    Teststepform,
    TestCasetestform,
    save_excelform
)
from project.models import CreateProject
from testcase.decorators import allowed_user
import openpyxl
from django.views.generic.base import TemplateView, View
from testcase.openpyxl_image_loader import SheetImageLoaderlist
from io import BytesIO
import base64
from testcase.models import (
    Teststep,
    TestCasetest,
    TagNewTestcase,
    save_excel)
from testplan.models import (
    TestPlan_test,
    TeststepforTestplan,
    TestCaseforTestplan)
import re
from difflib import SequenceMatcher
import json
import ast
from django.contrib.auth.models import User
from django.core.files.base import File
import os

'''
import excel
還沒有設定權限
'''
def str_remove(item):
    item = str(item)
    item2 = item.replace('[','')
    item2 = item2.replace(']','')
    return item2 
def im_2_b64(image):
    buff = BytesIO()
    image.save(buff, format="PNG")
    img_str = base64.b64encode(buff.getvalue())
    return img_str
'''
比較字的相似度
'''
def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

def debug_excel(name,data):
    description = data[0]
    condition = data[1]
    remark = data[2]
    with open("/res/testanswer.csv",'a') as fd:
        writer = csv.writer(fd)
        writer.writerow([i,skip,'手動模式處理 ',ddtimestart,ddtime,'',''])
    pass

def TestStepView(request,pk):    
    testcaseform = TestCasetestform()
    group_allow = ['creator']
    group = None
    group_len = None
    edit_clone_btn_cansee = False
    if request.user.groups.exists():
        print(request.user.groups.all())
        group_len = len(request.user.groups.all())
    print('group=',group,'allow=',group_allow)
    if group_len != None:
        i=0
        while i<group_len:
            group = request.user.groups.all()[i].name
            if group in group_allow:
                edit_clone_btn_cansee = True
            i+=1
    testcase = TestCasetest.objects.get(id=pk)
    teststep = Teststep.objects.filter(testcase = testcase)
    teststep_count = teststep.count()
    tag_object_list = testcase.tag.all()
    teststepform = Teststepform()
    teststepset = inlineformset_factory(TestCasetest, Teststep, fields=('number','description','condition','remark'), extra=0)
    form_teststep_list = teststepset(instance=testcase)
    testcase_information_html = [{'pk':testcase.pk,'name':testcase.name}]
    testcase_information_html = json.dumps(testcase_information_html)
    if 'save_teststep' in request.POST:
        '''
        teststep update
        '''
        Teststep.objects.filter(testcase=testcase).delete()
        teststep_dict = {}
        for item in request.POST:
            if 'id_teststep_set' in item:
                item_name = item.split('-')
                item_number = item_name[1]
                item_data = item_name[2]
                data = request.POST[item]
                try:
                    teststep_dict[item_number]
                    if item_data == 'description':
                        teststep_dict[item_number][0]=data
                    elif item_data == 'condition':
                        teststep_dict[item_number][1]=data
                    elif item_data == 'remark':
                        teststep_dict[item_number][2]=data
    
                except:
                    if item_data == 'description':
                        teststep_dict[item_number]=[data,None,None,None,None]
                    elif item_data == 'condition':
                        teststep_dict[item_number]=[None,data,None,None,None]
                    elif item_data == 'remark':
                        teststep_dict[item_number]=[None,None,data,None,None]
        count = 0
        for item in teststep_dict:
            if teststep_dict[item][0] != '':
                Teststep.objects.create(testcase=testcase,
                                        number=str(count+1),
                                        description=teststep_dict[item][0],
                                        condition=teststep_dict[item][1].replace('<p>&nbsp;</p>',''),
                                        remark=teststep_dict[item][2].replace('<p>&nbsp;</p>',''),)   
                count+=1
        tag_object_list = testcase.tag.all()
    elif 'save_info' in request.POST:
        detect_exist_name = False
        '''
        testcasename update
        '''
        testcasename = request.POST['testcasename']
        try:
            testcase_name_object = TestCasetest.objects.get(name=testcasename)
            if testcase_name_object.pk != testcase.pk:
                detect_exist_name = True
        except:
            pass
        if detect_exist_name == True:
            testcase_name_object = TestCasetest.objects.get(name=';;;;~~;`!!=-dsfasf, lsapdfisda,pfaisdfo ')
        else:
            pass
        testcase.name = testcasename
        testcase.description = request.POST['context']
        '''
        testcase update tag
        '''
        tag_list = request.POST.getlist('tag_list[]')
        for item in testcase.tag.all():
            testcase.tag.remove(item)
        for item in tag_list:
            try:
                testcase.tag.add(TagNewTestcase.objects.get(name=item))
            except:
                pass
        testcase.save()
    elif 'add_tag' in request.POST:
        bool_exists = False
        new_tag_name = request.POST['tag_name']
        try:
            TagNewTestcase.objects.get(name=new_tag_name)
            bool_exists = True
        except:
            TagNewTestcase.objects.create(name=new_tag_name)
        if bool_exists == True:
            TagNewTestcase.objects.get(name=';;;;~~;`!!=-dsfasf, lsapdfisda,pfaisdfo ')
    elif 'delete_testcase' in request.POST:
        testcase.delete()
        teststep.delete()
        return redirect('Dashboard')  
    
    elif 'clone' in request.POST:
        print('request',request.POST)
        if '-clone' in str(testcase.name):
            temp_clone_list = str(testcase.name).split('-clone')
            try:
                clone_num = int(temp_clone_list[-1])
                clonename = temp_clone_list[0]+'-clone'+str(clone_num+1)
            except:
                clonename = str(testcase.name)+'-clone1'
        else:
            clonename = str(testcase.name)+'-clone1'
        testcaseclone = TestCasetest.objects.create(name=clonename)
        testcaseclone = TestCasetest.objects.get(name = clonename)
        for instance in tag_object_list:
            testcaseclone.tag.add(TagNewTestcase.objects.get(name=instance.name))
        testcaseclone.save()
        for data in teststep:
            Teststep.objects.create(testcase=testcaseclone,
                                        number=str(data.number),
                                        description=data.description,
                                        condition=data.condition,
                                        remark=data.remark,)   
        return HttpResponseRedirect(reverse("teststep", args=[testcaseclone.pk]))
    tag_object = TagNewTestcase.objects.all()
    tag_list= []
    for item in tag_object_list:
        tag_list.append(item.name)
    tag_all_list = []
    for item in tag_object:
        tag_all_list.append(item.name)
    tag_list_json=json.dumps(tag_list)
    print('tag_list',tag_list)
    context = {'testcase':testcase,'teststep':teststep,'count':teststep_count,'form_teststep_list':form_teststep_list,'testcase_information_html':testcase_information_html,
               'tag_list':tag_list,'tag_list_json':tag_list_json,'tag_all_list':tag_all_list,'teststepform':teststepform,
               'edit_clone_btn_cansee':edit_clone_btn_cansee,'testcaseform':testcaseform}
    return render(request,'testcase/teststep.html',context)
    
# @allowed_user(allowed_roles=['creator'])
def TestcaseNew_test(request):
    form_testcase = TestCasetestform()
    teststepform = Teststepform()
    excel_form = save_excelform()
    '''
    更新teststep form
    '''
    if request.POST:
        print(request.POST)
        if 'save_testcase' in request.POST:
            bool_exists = False
            tag_list = request.POST.getlist('tag_list[]')
            print(tag_list)
            testcasename = request.POST['name']   
            content = request.POST['content']
            creator = request.user
            '''
            create testcase
            '''
            try:
                testcase_name_object = TestCasetest.objects.get(name=testcasename)
                bool_exists = True
            except:
                testcase = TestCasetest.objects.create(name=testcasename,description =content,creator=creator)
                testcase_name_object = TestCasetest.objects.get(name=testcasename)
                for item in tag_list:
                    print(item)
                    try:
                        tag= TagNewTestcase.objects.get(name=item)
                        print(tag.name)
                        testcase.tag.add(tag.id)
                    except Exception as e:
                        print('error = ',e,item)          
            if bool_exists == True:
                TestCasetest.objects.get(name=';;;;~~;`!!=-dsfasf, lsapdfisda,pfaisdfo ')
            '''
            teststep
            '''
            teststep_dict = {}
            for item in request.POST:
                if 'id_teststep_set' in item:
                    item_name = item.split('-')
                    item_number = item_name[1]
                    item_data = item_name[2]
                    data = request.POST[item]
                    try:
                        teststep_dict[item_number]
                        if item_data == 'description':
                            teststep_dict[item_number][0]=data
                        elif item_data == 'condition':
                            teststep_dict[item_number][1]=data
                        elif item_data == 'remark':
                            teststep_dict[item_number][2]=data
                    except:
                        if item_data == 'description':
                            teststep_dict[item_number]=[data,None,None,None]
                        elif item_data == 'condition':
                            teststep_dict[item_number]=[None,data,None,None]
                        elif item_data == 'remark':
                            teststep_dict[item_number]=[None,None,data,None]
            count = 0
            for item in teststep_dict:
                if teststep_dict[item][0] != '':
                    Teststep.objects.create(testcase=testcase_name_object,
                                            number=str(count+1),
                                            description=teststep_dict[item][0],
                                            condition=teststep_dict[item][1].replace('<br />','').replace('<p>&nbsp;</p>','').replace('</br>',''),
                                            remark=teststep_dict[item][2].replace('<br />','').replace('<p>&nbsp;</p>','').replace('</br>',''))   
                    count+=1
            return JsonResponse({'pk':str(testcase_name_object.id)})
        elif 'add_tag' in request.POST:
            bool_exists = False
            new_tag_name = request.POST['tag_name']
            try:
                TagNewTestcase.objects.get(name=new_tag_name)
                bool_exists = True
            except:
                TagNewTestcase.objects.create(name=new_tag_name)
            response_data = {}
            response_data['result'] = 'error'
            response_data['message'] = 'Some error message'
            if bool_exists == True:
                TagNewTestcase.objects.get(name=';;;;~~;`!!=-dsfasf, lsapdfisda,pfaisdfo ')
        elif 'upload_testcase_btn' in request.POST:
            form = save_excelform(request.POST, request.FILES)
            if form.is_valid():
                newdoc = save_excel(filename = request.FILES['file'].name ,file = request.FILES['file'])
                newdoc.save()
                print(newdoc)
                print('path=',newdoc.file.path)
                file_path = newdoc.file.path
                wb = openpyxl.load_workbook(file_path)
                print(wb.sheetnames)
                sheet_list = wb.sheetnames
                file_name_testcase = newdoc.filename
                file_name_testcase = file_name_testcase.replace('.xlsx','')
                file_name_testcase = file_name_testcase.replace('.xls','')
                testcase_name = file_name_testcase
                print('testcasename = ',testcase_name)
                return JsonResponse({'sheet_list':sheet_list,'filepath':file_path,'testcase_name':testcase_name})
        elif 'get_sheetname' in request.POST:
            file_path = request.POST['filepath']
            '''
            設定標準的list
            '''
            title_list_title = [['編號','number'],['驗 證 項 目','description'],['驗 證 條 件','condition'],['備註','remark']]
            sheet_name = request.POST['select_sheet']
            title_list = [] 
            c='ABCDEFGHIJKLMNOPQRSTUVWXYZ'
            count_header = 0
            while title_list == [] or len(title_list)<4:
                df = pd.read_excel(file_path, sheet_name=sheet_name , header=count_header)
                count=0
                for item in df:
                    if 'Unnamed' not in item:
                        title_list.append([c[count],item])
                    count+=1
                if title_list == [] or len(title_list)<4:
                    count_header+=1
            '''
            將一樣的數字放在option 的list裡面
            '''
            count = 0 
            for data in title_list_title:
                score = 0
                for data2 in title_list:
                    print('data2',data2,'title_list',title_list)
                    if score < similar(data[0],data2[1]):
                        score = similar(data[0],data2[1])
                        ans = data2[1]
                title_list_title[count].append(ans)
                count += 1
            return JsonResponse({'title_list_title':title_list_title,'title_list':title_list,'sheet_name':sheet_name,'count_header':count_header})
        elif 'save_data' in request.POST:
            file_path = request.POST['filepath']
            '''
            選取的標題名稱
            '''
            number_title =request.POST.getlist('number[]')
            description_title = request.POST.getlist('description[]')
            condition_title = request.POST.getlist('condition[]')
            remark_title = request.POST.getlist('remark[]')
            usecols_title_list = [number_title[0],description_title[0],condition_title[0],remark_title[0]]
            usecols_title_list_name = [number_title[1],description_title[1],condition_title[1],remark_title[1]]
            count_header = int(request.POST['count_header'])
            sheetname = request.POST['sheet_name'] 
            testcase_name = request.POST['testcase_name']
            df = pd.read_excel(file_path,sheet_name=sheetname, header=count_header,usecols=usecols_title_list_name)
            i=0
            data_list={}
            count=1
            while i<len(df):
                temp_data = (df[i:i+1].values).tolist()
                temp_data = temp_data[0]
                # print(temp_data)
                if str(temp_data[0])!='nan': 
                    number=str(temp_data[0]).replace('.0','')
                    if str(temp_data[1])!='nan':
                        description = temp_data[1]
                    else:
                        description = ''
                    if str(temp_data[2]) != 'nan':
                        condition = temp_data[2]
                    else:
                        condition = ''
                    if str(temp_data[3]) != 'nan':
                        remark = temp_data[3]
                    else:
                        remark = ''   
                    data_list[count]=[i,description,condition,remark]
                    count+=1
                i+=1
            key_list = list(data_list)
            '''
            引入圖片
            '''
            #loading the Excel File and the sheet
            pxl_doc = openpyxl.load_workbook(file_path)
            sheet = pxl_doc[sheetname]
            #calling the image_loader
            image_loader = SheetImageLoaderlist(sheet)
            image_dict = {}     #記錄圖片的內容
            for target in image_loader._images:
                image_list = image_loader.get(target)
                for image in image_list:
                    img_b64 = im_2_b64(image)
                    try:
                        image_dict[target]
                        if str(img_b64).replace("b'",'').replace("'","") not in image_dict[target]:
                            image_dict[target].append(str(img_b64).replace("b'",'').replace("'",""))
                    except:
                        image_dict[target] = [str(img_b64).replace("b'",'').replace("'","")]
            image_list= []      # 記錄有圖片的位子
            for item in image_dict:
                a = re.findall(r'\d+',item)
                a = str(a).replace('[','').replace(']','').replace("'","")
                temp_item = item.replace(a,'')
                if temp_item in usecols_title_list:
                    image_list.append(item)
            '''
            將圖片加入
            '''
            for item in image_list:
                a = re.findall(r'\d+',item)
                a = str(a).replace('[','').replace(']','').replace("'","")
                number = int(a)
                pic_input = ''
                for image_base64 in image_dict[item]:
                    pic_input += '<p><img src="data:image/png;base64,'+image_base64+'"/></p>'
                for data in data_list:
                    if int(data_list[data][0])>number:
                        last_key = key_list[key_list.index(data)-1]
                        if usecols_title_list[1] in item:
                            data_list[last_key][1]=data_list[last_key][1]+pic_input
                        elif usecols_title_list[2] in item:
                            data_list[last_key][2]=data_list[last_key][2]+pic_input
                        elif usecols_title_list[3] in item:
                            data_list[last_key][3]=data_list[last_key][3]+pic_input
                        break
            # print(data_list)
            os.remove(file_path)
            save_excel.objects.all().delete()
            
            return JsonResponse({'data_list':data_list})



    context ={'form_testcase':form_testcase,'teststepform':teststepform,'excel_form':excel_form}
    return render(request,'testcase/create_testcase.html',context)