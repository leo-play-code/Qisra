from django.contrib import messages
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.urls import reverse
from django.utils.translation import gettext_lazy as _
import pandas as pd
from django.forms import inlineformset_factory
from testcase.forms import (
    Teststepform,
    TestCasetestform,
    save_excelform,
)
from project.models import Project
from testcase.decorators import allowed_user
import openpyxl
from django.views.generic.base import TemplateView, View
from testcase.openpyxl_image_loader import SheetImageLoaderlist
from io import BytesIO
import base64
from testcase.models import (
    Teststep,
    Testcase,
    Tag,testcase_file,teststep_file,
    testcase_import_excel)

from testplan.models import Testrun


import re
from difflib import SequenceMatcher
import json
import ast
from django.contrib.auth.models import User
from django.core.files.base import File
import os
from datetime import datetime as dt2
import datetime as dt
import csv
from django.contrib.auth.decorators import login_required
from signup_app.decorators import allowed_user
from simple_history.utils import update_change_reason
import time
# clone method
from main.clone_method import Clone_Models

'''
import excel
還沒有設定權限
'''
@login_required
def add_tag(request):
    bool_exists = False
    new_tag_name = request.GET['tag_name']
    try:
        Tag.objects.get(name=new_tag_name)
        bool_exists = True
    except:
        new_tag = Tag.objects.create(name=new_tag_name)
        return JsonResponse({'name':new_tag_name,'id':str(new_tag.id),'status':'pass'})
    if bool_exists == True:
        return JsonResponse({'name':new_tag_name,'status':'fail'})

def str_remove(item):
    item = str(item)
    item2 = item.replace('[', '')
    item2 = item2.replace(']', '')
    return item2

def im_2_b64(image):
    buff = BytesIO()
    image.save(buff, format="PNG")
    img_str = base64.b64encode(buff.getvalue())
    return img_str


'''
比較字的相似度
'''


def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

def get_history(history_model):
    '''
    history_model (Model.class): model.history.all()
    '''
    history_dict = {}
    for h in history_model:
        try:
            history_reason = (h.history_change_reason).split('||')
            count = 0
            history_reason_list_all = []
            history_reason_list = []
            for reason_item in history_reason:
                if count < 4:
                    history_reason_list.append(reason_item)
                else:
                    history_reason_list_all.append(history_reason_list)
                    history_reason_list = [reason_item]
                    count = 0
                count += 1
            history_dict[h] = history_reason_list_all
        except Exception as e:
            pass
    return history_dict

def get_file_data(model):
    '''
    model : file model
    '''
    format_data = "%Y-%m-%d %H:%M:%S"
    file_path = model.file.path
    try:
        file_path = file_path.split('media/')[1]
    except:
        file_path = file_path.split('media\\')[1]
    filename = model.filename
    temp_time = str(model.date_created)
    temp_time = temp_time.split(' ')
    temp_date = temp_time[0]
    temp_time = (temp_time[1].split('.'))[0]
    temp_time = temp_date+' '+temp_time
    temp_time = dt2.strptime(temp_time, format_data)
    temp_time = temp_time + dt.timedelta(hours=8)
    temp_time = str(temp_time)
    return file_path,filename,temp_time



@login_required
def TestStepView(request, pk):
    testcase = Testcase.objects.get(id=pk)
    tag_object_list = testcase.tag.all()
    testcaseform = TestCasetestform()
    testcase_file_list = testcase_file.objects.filter(testcase = testcase)
    teststep = Teststep.objects.filter(testcase=testcase).order_by('number')
    teststep_orignal_id_list = []
    for item in teststep:
        teststep_orignal_id_list.append(int(item.id))
    teststepform = Teststepform()
    teststepset = inlineformset_factory(Testcase, Teststep, fields=(
        'number', 'description', 'condition', 'remark'), extra=0)
    form_teststep_list = teststepset(instance=testcase,queryset=Teststep.objects.order_by('number'))
    current_user = request.user   
    # teststep file
    teststep_dict = {}
    for item in teststep:
        teststep_dict[item] = teststep_file.objects.filter(teststep=item)
    # get history dict
    testcase_history_model = testcase.history.all()
    history_dict = get_history(testcase_history_model)
    # reason_history
    temp_blank = ''
    # detect testcase in any testplan-group testrun or not 
    testrun_count = Testrun.objects.filter(testcase = testcase).count()
    if 'save_teststep' in request.POST:
        '''
        teststep update
        '''
        change_reason = ''
        for item in request.POST:
            if 'id_teststep_set' in item:
                item_name = item.split('-')
                item_testsstep_id = item_name[1]
                item_data = request.POST.getlist(item)
                try:
                    temp_teststep = Teststep.objects.get(id=int(item_testsstep_id))
                    temp_teststep.number = int(item_data[0])
                    temp_teststep.description = item_data[1].replace('<p></p>','')
                    temp_teststep.condition = item_data[2].replace('<p></p>','')
                    temp_teststep.remark = item_data[3].replace('<p></p>','')
                    temp_teststep.save()
                    # history
                    new_record, old_record, *_ = temp_teststep.history.all()
                    delta = new_record.diff_against(old_record)
                    for change in delta.changes:
                        if change.field != 'number':
                            if change.field == 'description':
                                title = '驗證項目'
                            elif change.field == 'condition':
                                title = '驗證條件'
                            elif change.field == 'remark':
                                title = '備註'
                            if len(change.old) > len(change.new):
                                try:
                                    res = ''.join(change.old.split(change.new))
                                except:
                                    res = change.old
                            else:
                                try:
                                    res = ''.join(change.new.split(change.old))
                                except:
                                    res = change.new
                            final_data = res.strip().replace('<br>', '').replace('<br />', '').replace(' ', '')
                            final_data = final_data.replace('<p></p>','')
                            if final_data != '':
                                change_reason += "編輯||編號{}-{}:\n||{}||{}||".format(
                                    str(item_data[0]), title, change.old, change.new)
                    teststep_orignal_id_list.remove(int(item_testsstep_id))
                except:
                    Teststep.objects.create(testcase=testcase,
                                            number= int(item_data[0]),
                                            description=item_data[1],
                                            condition=item_data[2],
                                            remark=item_data[3])
                    change_reason += "新增||編號{}:\n||{}||{}||".format(
                        str(item_data[0]), temp_blank, item_data[1])
        # print('teststep_orignal_id_list',teststep_orignal_id_list)
        for item in teststep_orignal_id_list:
            try:
                temp_teststep = Teststep.objects.get(id=item)
                change_reason += "刪除||編號{}:\n||{}||{}||".format(
                        str(temp_teststep.number), temp_teststep.description,temp_blank)
                temp_teststep.delete()
            except:
                pass
            
        testcase.save()
        if change_reason != '':
            update_change_reason(testcase, change_reason)
        teststep = Teststep.objects.filter(testcase=testcase).order_by('number')
        
        # teststep file
        teststep_dict = {}
        for item in teststep:
            teststep_dict[str(item.number)] = {
                'id':item.id,
                'description':item.description,
                'condition':item.condition,
                'remark':item.remark
            }
            temp_list = []
            for file_item in teststep_file.objects.filter(teststep=item):
                format_data = "%Y-%m-%d %H:%M:%S"
                file_path = file_item.file.path
                file_path = file_path.split('media/')[1]
                filename = file_item.filename
                temp_time = str(file_item.date_created)
                temp_time = temp_time.split(' ')
                temp_date = temp_time[0]
                temp_time = (temp_time[1].split('.'))[0]
                temp_time = temp_date+' '+temp_time
                temp_time = dt2.strptime(temp_time, format_data)
                temp_time = temp_time + dt.timedelta(hours=8)
                temp_time = str(temp_time)
                temp_list.append([filename,file_path,file_item.uploader.username,temp_time,str(file_item.id)])
            teststep_dict[str(item.number)]['filelist']= temp_list
        return JsonResponse({
            'teststep_dict':teststep_dict
        })
    elif 'save_info' in request.POST:
        detect_exist_name = False
        '''
        testcasename update
        '''
        testcasename = request.POST['testcasename']
        try:
            testcase_name_object = Testcase.objects.get(name=testcasename)
            if testcase_name_object.pk != testcase.pk:
                detect_exist_name = True
        except:
            pass
        if detect_exist_name == True:
            testcase_name_object = Testcase.objects.get(
                name=';;;;~~;`!!=-dsfasf, lsapdfisda,pfaisdfo ')
        else:
            pass
        testcase.name = testcasename

        '''
        testcase update tag
        '''
        tag_list = request.POST.getlist('tag_list[]')
        orignal_tag_list = []
        for item in testcase.tag.all():
            orignal_tag_list.append(item.name)
        tag_list_str = ''
        for item in tag_list:
            tag_list_str += (item+' ')
        orignal_tag_list_str = ''
        for item in orignal_tag_list:
            orignal_tag_list_str += (item+' ')
        change_reason = ''
        s = set(orignal_tag_list)
        need_add = [x for x in tag_list if x not in s]
        s = set(tag_list)
        need_delete = [x for x in orignal_tag_list if x not in s]
        if len(need_delete) > 0 or len(need_add) > 0:
            title = '標籤'
            if orignal_tag_list_str.replace(' ', '') != tag_list_str.replace(' ', ''):
                change_reason += "編輯|| {}:|| {}||{}||".format(
                    title, orignal_tag_list_str, tag_list_str)
        for item in need_delete:
            testcase.tag.remove(Tag.objects.get(name=item))
        for item in need_add:
            try:
                testcase.tag.add(Tag.objects.get(name=item))
            except:
                pass
        testcase.save()
        new_record, old_record, *_ = testcase.history.all()
        delta = new_record.diff_against(old_record)
        for change in delta.changes:
            if change.field == 'name':
                title = '名稱'
            change_reason += "編輯|| {}:\n|| {}||{}||".format(
                title, change.old, change.new)
        if change_reason != '':
            update_change_reason(testcase, change_reason)
    elif 'save_context_data' in request.POST:
        testcase.description = request.POST['context']
        testcase.save()
        change_reason = ''
        new_record, old_record, *_ = testcase.history.all()
        delta = new_record.diff_against(old_record)
        for change in delta.changes:
            if change.field == 'description':
                title = '描述'
            change_reason += "編輯|| {}:\n || {}|| {} ||".format(
                title, change.old, change.new)
        if change_reason != '':
            update_change_reason(testcase, change_reason)
    elif 'delete_verify' in request.POST:
        testcase.delete()
        return redirect('Dashboard')
    elif 'upload_file_to_testcase' in request.POST:
        newdoc = testcase_file(
            filename=request.FILES['file'].name, file=request.FILES['file'],testcase = testcase,uploader=current_user)
        newdoc.save()
        # file path
        file_path,filename,temp_time=get_file_data(newdoc)
        # history
        change_reason = ''
        change_reason += "添加|| {}:\n || {}|| {} ||".format(
            '附件',temp_blank, newdoc.filename)
        testcase.save()
        if change_reason != '':
            update_change_reason(testcase, change_reason)
        return JsonResponse({'id':str(newdoc.id),'path':file_path,'filename':filename,'upload_time':temp_time,'uploader':newdoc.uploader.username})
    elif 'delete_upload_file' in request.POST:
        delete_id = request.POST['delete_id']
        newdoc = testcase_file.objects.get(id=int(delete_id))
        filename = newdoc.filename
        file_path = newdoc.file.path
        newdoc.delete()
        # os.remove(file_path)
        # history
        change_reason = ''
        change_reason += "刪除|| {}:\n || {}|| {} ||".format(
            '附件',filename, temp_blank)
        testcase.save()
        if change_reason != '':
            update_change_reason(testcase, change_reason)
    elif 'upload_file_to_teststep' in request.POST:
        temp_id = request.POST['id']
        temp_teststep = Teststep.objects.get(id=temp_id)
        newdoc = teststep_file(
            filename=request.FILES['file'].name, file=request.FILES['file'],teststep = temp_teststep,uploader=current_user)
        newdoc.save()
        file_path,filename,temp_time=get_file_data(newdoc)
        # history
        change_reason = ''
        change_reason += "添加附件|| 編號-{}:\n || {}|| {} ||".format(
            str(temp_teststep.number),temp_blank, newdoc.filename)
        testcase.save()
        if change_reason != '':
            update_change_reason(testcase, change_reason)
        return JsonResponse({'id':str(newdoc.id),'path':file_path,'filename':filename,'upload_time':temp_time,'uploader':newdoc.uploader.username})
    elif 'delete_teststep_file' in request.POST:
        delete_id = request.POST['delete_id']
        newdoc = teststep_file.objects.get(id=int(delete_id))
        temp_teststep = newdoc.teststep
        filename = newdoc.filename
        file_path = newdoc.file.path
        newdoc.delete()
        # os.remove(file_path)
        # history
        change_reason = ''
        change_reason += "刪除附件|| 編號-{}:\n || {}|| {} ||".format(
            temp_teststep.number,filename, temp_blank)
        testcase.save()
        if change_reason != '':
            update_change_reason(testcase, change_reason)
    elif 'clone' in request.POST:
        testcaseclone = Clone_Models(request.user,'testcase',testcase)
        return HttpResponseRedirect(reverse("teststep", args=[testcaseclone.pk]))
    tag_object = Tag.objects.all()
    tag_list = []
    for item in tag_object_list:
        tag_list.append(item.name)
    tag_all_list = []
    for item in tag_object:
        tag_all_list.append(item.name)
    tag_list_json = json.dumps(tag_list)
    context = {'testcase': testcase, 'teststep_dict': teststep_dict, 'form_teststep_list': form_teststep_list, 
               'tag_list': tag_list, 'tag_list_json': tag_list_json, 'tag_all_list': tag_all_list, 'teststepform': teststepform,'testcase_file_list':testcase_file_list,
               'testcaseform': testcaseform, 'history_dict': history_dict,'testrun_count':testrun_count}
    return render(request, 'testcase/teststep_view.html', context)


@login_required
@allowed_user(allowed_roles=['creator'])
def TestcaseNew_test(request):
    form_testcase = TestCasetestform()
    teststepform = Teststepform()
    excel_form = save_excelform()
    '''
    更新teststep form
    '''
    if request.POST:
        if 'save_testcase' in request.POST:
            bool_exists = False
            tag_list = request.POST.getlist('tag_list[]')
            testcasename = request.POST['name']
            content = request.POST['content']
            creator = request.user
            '''
            create testcase
            '''
            try:
                testcase_name_object = Testcase.objects.get(
                    name=testcasename)
                bool_exists = True
            except:
                testcase = Testcase.objects.create(
                    name=testcasename, description=content, creator=creator)
                testcase_name_object = Testcase.objects.get(
                    name=testcasename)
                for item in tag_list:
                    try:
                        tag = Tag.objects.get(name=item)
                        testcase.tag.add(tag.id)
                    except Exception as e:
                        print('error = ', e, item)
            if bool_exists == True:
                Testcase.objects.get(
                    name=';;;;~~;`!!=-dsfasf, lsapdfisda,pfaisdfo ')

            '''
            teststep
            '''
            teststep_dict = {}
            for item in request.POST:
                if 'id_teststep_set' in item:
                    item_name = item.split('-')
                    item_number = item_name[1]
                    item_data = item_name[2]
                    data = request.POST[item]
                    try:
                        teststep_dict[item_number]
                        if item_data == 'description':
                            teststep_dict[item_number][0] = data
                        elif item_data == 'condition':
                            teststep_dict[item_number][1] = data
                        elif item_data == 'remark':
                            teststep_dict[item_number][2] = data

                    except:
                        if item_data == 'description':
                            teststep_dict[item_number] = [
                                data, None, None, None]
                        elif item_data == 'condition':
                            teststep_dict[item_number] = [
                                None, data, None, None]
                        elif item_data == 'remark':
                            teststep_dict[item_number] = [
                                None, None, data, None]

            count = 0
            for item in teststep_dict:
                if teststep_dict[item][0] != '':
                    Teststep.objects.create(testcase=testcase_name_object,
                                            number=int(count+1),
                                            description=teststep_dict[item][0],
                                            condition=teststep_dict[item][1],
                                            remark=teststep_dict[item][2],
                                            )
                    count += 1
            return JsonResponse({'pk': str(testcase_name_object.id)})
        elif 'upload_testcase_btn' in request.POST:
            form = save_excelform(request.POST, request.FILES)
            if form.is_valid():
                newdoc = testcase_import_excel(
                    filename=request.FILES['file'].name, file=request.FILES['file'])
                newdoc.save()
                file_path = newdoc.file.path
                wb = openpyxl.load_workbook(file_path)
                sheet_list = wb.sheetnames
                file_name_testcase = newdoc.filename
                file_name_testcase = file_name_testcase.replace('.xlsx', '')
                file_name_testcase = file_name_testcase.replace('.xls', '')
                testcase_name = file_name_testcase
                return JsonResponse({'sheet_list': sheet_list, 'filepath': file_path, 'testcase_name': testcase_name})
        elif 'get_sheetname' in request.POST:
            file_path = request.POST['filepath']
            '''
            設定標準的list
            '''
            title_list_title = [['編號', 'number'], ['驗 證 項 目', 'description'], [
                '驗 證 條 件', 'condition'], ['備註', 'remark']]
            sheet_name = request.POST['select_sheet']
            title_list = []
            c = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
            count_header = 0
            while title_list == [] or len(title_list) < 4:
                df = pd.read_excel(
                    file_path, sheet_name=sheet_name, header=count_header)
                count = 0
                for item in df:
                    if 'Unnamed' not in item:
                        title_list.append([c[count], item])
                    count += 1
                if title_list == [] or len(title_list) < 4:
                    count_header += 1
            '''
            將一樣的數字放在option 的list裡面
            '''
            count = 0
            for data in title_list_title:
                score = 0
                for data2 in title_list:
                    if score < similar(data[0], data2[1]):
                        score = similar(data[0], data2[1])
                        ans = data2[1]
                title_list_title[count].append(ans)
                count += 1
            return JsonResponse({'title_list_title': title_list_title, 'title_list': title_list, 'sheet_name': sheet_name, 'count_header': count_header})
        elif 'save_data' in request.POST:
            file_path = request.POST['filepath']
            '''
            選取的標題名稱
            '''
            number_title = request.POST.getlist('number[]')
            description_title = request.POST.getlist('description[]')
            condition_title = request.POST.getlist('condition[]')
            remark_title = request.POST.getlist('remark[]')
            usecols_title_list = [
                number_title[0], description_title[0], condition_title[0], remark_title[0]]
            usecols_title_list_name = [
                number_title[1], description_title[1], condition_title[1], remark_title[1]]
            count_header = int(request.POST['count_header'])
            sheetname = request.POST['sheet_name']
            testcase_name = request.POST['testcase_name']
            df = pd.read_excel(file_path, sheet_name=sheetname,
                               header=count_header, usecols=usecols_title_list_name)
            i = 0
            data_list = {}
            count = 1
            while i < len(df):
                temp_data = (df[i:i+1].values).tolist()
                temp_data = temp_data[0]
                if str(temp_data[0]) != 'nan':
                    number = str(temp_data[0]).replace('.0', '')
                    if str(temp_data[1]) != 'nan':
                        description = temp_data[1]
                    else:
                        description = ''
                    if str(temp_data[2]) != 'nan':
                        condition = temp_data[2]
                    else:
                        condition = ''
                    if str(temp_data[3]) != 'nan':
                        remark = temp_data[3]
                    else:
                        remark = ''
                    try:
                        data_list[count-1][0] = i+1
                        i += 1
                    except:
                        pass
                    data_list[count] = [i, description, condition, remark]
                    count += 1
                i += 1
            try:
                data_list[count-1][0] = i+1
            except:
                pass
            key_list = list(data_list)
            '''
            引入圖片
            '''
            # loading the Excel File and the sheet
            pxl_doc = openpyxl.load_workbook(file_path)
            sheet = pxl_doc[sheetname]
            # calling the image_loader
            image_loader = SheetImageLoaderlist(sheet)
            image_dict = {}  # 記錄圖片的內容
            for target in image_loader._images:
                image_list = image_loader.get(target)
                for image in image_list:
                    img_b64 = im_2_b64(image)
                    try:
                        image_dict[target]
                        if str(img_b64).replace("b'", '').replace("'", "") not in image_dict[target]:
                            image_dict[target].append(
                                str(img_b64).replace("b'", '').replace("'", ""))
                    except:
                        image_dict[target] = [
                            str(img_b64).replace("b'", '').replace("'", "")]
            image_list = []      # 記錄有圖片的位子
            for item in image_dict:
                a = re.findall(r'\d+', item)
                a = str(a).replace('[', '').replace(']', '').replace("'", "")
                temp_item = item.replace(a, '')
                if temp_item in usecols_title_list:
                    image_list.append(item)
            '''
            將圖片加入
            '''
            for item in image_list:
                a = re.findall(r'\d+', item)
                a = str(a).replace('[', '').replace(']', '').replace("'", "")
                number = int(a)
                pic_input = ''
                for image_base64 in image_dict[item]:
                    pic_input += '<p><img src="data:image/png;base64,'+image_base64+'"/></p>'
                for data in data_list:
                    if int(data_list[data][0]) >= number:
                        last_key = key_list[key_list.index(data)]
                        if usecols_title_list[1] in item:
                            data_list[last_key][1] = data_list[last_key][1]+pic_input
                        elif usecols_title_list[2] in item:
                            data_list[last_key][2] = data_list[last_key][2]+pic_input
                        elif usecols_title_list[3] in item:
                            data_list[last_key][3] = data_list[last_key][3]+pic_input
                        elif usecols_title_list[4] in item:
                            data_list[last_key][4] = data_list[last_key][4]+pic_input
                        break
            os.remove(file_path)
            testcase_import_excel.objects.all().delete()

            return JsonResponse({'data_list': data_list})

    context = {'form_testcase': form_testcase,
               'teststepform': teststepform, 'excel_form': excel_form}
    return render(request, 'testcase/create_testcase.html', context)
